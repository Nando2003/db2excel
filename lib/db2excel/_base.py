from abc import ABC, abstractmethod

from typing import List, Tuple

from sqlalchemy.engine.base import Engine
from sqlalchemy import inspect, text

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .exceptions import InvalidPathException
from os.path import abspath, isdir, isfile, join
from os import remove, getcwd

class DatabaseToExcel(ABC):
    
    def __init__(
        self,
        download_path    : str  = None,
        excel_name       : str  = None,
        overwrite        : bool = False,
    ):
        self.download_path = download_path
        self.excel_name    = excel_name or 'output'
        self.overwrite     = overwrite
        
        self._process()
        
    def _process(self) -> None:
        self.validate_download_path()
        engine = self.creating_engine()
        tables = self.get_tables_from_db(engine)
        
        for table in tables:
            columns = self.get_columns_from_table(engine, table)
            query   = self.search_all_the_data(columns, table)
            data    = self.get_data_from_table(engine, query)
            
            self.list_to_sheet(
                table_name = table,
                columns    = columns,
                data       = data
            )
    
    def validate_download_path(self) -> None:
        if not self.download_path:
            self.download_path = getcwd()
            
        self.download_path = abspath(self.download_path)
        
        if isdir(self.download_path) is False:
            raise InvalidPathException(f'The specified path "{self.download_path}" does not exist')
        
        self.excel_name = self.excel_name + ".xlsx"
        self.download_path = join(self.download_path, self.excel_name)
        
        if isfile(self.download_path):
            if self.overwrite is True:
                remove(self.download_path)
            else:
                raise FileExistsError(f'The specified "{self.excel_name}" file name exists')
        
    @abstractmethod
    def creating_engine(self) -> Engine:
        pass
    
    def get_tables_from_db(self, engine:Engine) -> list:
        return inspect(engine).get_table_names()
    
    def get_columns_from_table(self, engine:Engine, table_name:str) -> list:
        columns = inspect(engine).get_columns(table_name)
        return [column['name'] for column in columns]
    
    @abstractmethod
    def search_all_the_data(self, columns:list, table_name:str) -> str:
        pass
    
    def get_data_from_table(self, engine:Engine, query:str) -> List[Tuple]:
        with engine.connect() as connection:
            result = connection.execute(query)
            return [tuple(data) for data in result]
        
    def list_to_sheet(self, table_name:str, columns:list, data:List[Tuple]) -> None:
        try:
            workbook = load_workbook(self.download_path)
        except FileNotFoundError:
            workbook = Workbook()
        
        if "Sheet" in workbook.sheetnames:
            worksheet = workbook["Sheet"]
            worksheet.title = table_name
        else:
            worksheet = workbook.create_sheet(
                title=table_name
            )
        
        bold_font = Font(bold=True)
        for i, column in enumerate(columns):
            cell = worksheet.cell(
                row=1,
                column=i+1,
                value=column
            )
            cell.font = bold_font
        
        for row in data:
            worksheet.append(row)
        
        workbook.save(self.download_path) 
    
    def __str__(self) -> str:
        return f"Creating an Excel file: {self.excel_name}"
    
    
    