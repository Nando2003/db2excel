from sqlalchemy import inspect, text
from sqlalchemy.engine.base import Engine
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from typing import Union, List, Tuple
import os

from .exceptions import InvalidPathException

class DatabaseToExcel:
    
    def __init__(
        self,
        database       :str  = None,
        username       :str  = None,
        password       :str  = None,
        host           :str  = None,
        port:Union[str, int] = None,
        download_path  :str  = None,
        excel_name     :str  = None,
        overwrite      :bool = False
    ):
        self.database      = database
        self.username      = username
        self.password      = password
        self.host          = host
        self.port          = str(port) if isinstance(port, int) else port
        self.download_path = download_path
        self.excel_name    = excel_name or 'output'
        self.overwrite     = overwrite
        
        self._process()
        
    def _process(self) -> None:
        self.download_path_validate()
        engine = self.creating_engine()
        tables = self.get_tables_from_db(engine)
        
        for table in tables:
            columns = self.get_columns_from_table(engine, table)
            data    = self.get_data_from_table(engine, columns, table)
            
            self.list_to_sheet(
                table_name =table,
                columns    =columns,
                data       =data
            )

    def download_path_validate(self) -> None:
        if not os.path.isdir(self.download_path):
            raise InvalidPathException(f'The specified path "{self.download_path}" does not exist')
        
        self.excel_name = self.excel_name + ".xlsx"
        self.download_path = os.path.join(self.download_path, self.excel_name)
        
        if os.path.isfile(self.download_path) and not self.overwrite:
            raise FileExistsError(f'The specified "{self.excel_name}" file name exists')

    def list_to_sheet(self, table_name:str, columns:List[str], data:List[Tuple]) -> None:
        try:
            workbook = load_workbook(self.download_path)
            
        except FileNotFoundError:
            workbook = Workbook()
        
        if table_name in workbook.sheetnames:
            worksheet = workbook[table_name]
        else:
            worksheet = workbook.create_sheet(title=table_name)
        
        bold_font = Font(bold=True)
        for i, column in enumerate(columns):
            cell = worksheet.cell(row=1, column=i+1, value=column)
            cell.font = bold_font
        
        for row in data:
            worksheet.append(row)
        
        workbook.save(self.download_path)

    def creating_engine(self) -> Engine:
        raise NotImplementedError("Subclasses must implement this method")

    def get_tables_from_db(self, engine:Engine) -> List[str]:
        return inspect(engine).get_table_names()

    def get_columns_from_table(self, engine:Engine, table_name:str) -> List[str]:
        columns = inspect(engine).get_columns(table_name)
        return [column['name'] for column in columns]

    def get_data_from_table(self, engine:Engine, columns:List[str], table_name:str) -> List[Tuple]:
        columns_name = ', '.join(f'"{col}"' for col in columns)
        query = text(f'SELECT {columns_name} FROM "{table_name}"')
        with engine.connect() as connection:
            result = connection.execute(query)
            return [tuple(row) for row in result]
