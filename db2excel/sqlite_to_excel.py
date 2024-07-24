from sqlalchemy import create_engine, inspect, text
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from typing import List, Tuple
import os

from exceptions import InvalidPathException

class SqliteToExcel:
    def __init__(self, db_path:str, download_path:str, excel_name:str= None, overwrite:bool= False) -> None:
        self.download_path = download_path
        self.excel_name    = excel_name or 'output'
        self.db_path       = db_path
        self.overwrite     = overwrite
        self._process()
    
    def _process(self) -> None:
        self.download_path_validate()
        engine = self.connecting_to_db()
        tables = self.get_tables_from_db(engine)
        
        for table in tables:
            columns = self.get_columns_from_table(engine, table)
            data    = self.get_data_from_table(engine, columns, table)
            
            self.list_to_sheet(
                table_name=table,
                columns=columns,
                data=data
            )
            
    def connecting_to_db(self) -> create_engine:
        return create_engine(f'sqlite:///{self.db_path}')
    
    def get_tables_from_db(self, engine:create_engine) -> list:
        return inspect(engine).get_table_names()
    
    def get_columns_from_table(self, engine:create_engine, table_name:str) -> list:
        columns = inspect(engine).get_columns(table_name)
        return [column['name'] for column in columns]
    
    def get_data_from_table(self, engine:create_engine, columns:list, table_name:str) -> List[Tuple]:
        columns_name = ', '.join(columns)
        query = text(f'SELECT {columns_name} FROM "{table_name}"')
        with engine.connect() as connection:
            result = connection.execute(query)
            return [tuple(data) for data in result]
    
    def download_path_validate(self) -> None:
        if os.path.isdir(self.download_path) is False:
            raise InvalidPathException(f'The specified {self.download_path} does not exist')
        
        self.excel_name = self.excel_name + ".xlsx"
        self.download_path = os.path.join(self.download_path, self.excel_name)
        
        if os.path.isfile(self.download_path) is True and not(self.overwrite):
            raise FileExistsError(f'The specified {self.excel_name} file exists')
    
    def list_to_sheet(self, table_name:str, columns:list, data:List[Tuple]) -> None:
        try:
            workbook = load_workbook(self.download_path)
        except FileNotFoundError:
            workbook = Workbook()
        
        if "Sheet" in workbook.sheetnames:
            worksheet = workbook["Sheet"]
            worksheet.title = table_name
        else:
            worksheet = workbook.create_sheet(
                title=table_name
            )
        
        bold_font = Font(bold=True)
        for i, column in enumerate(columns):
            cell = worksheet.cell(
                row=1,
                column=i+1,
                value=column
            )
            cell.font = bold_font
        
        for row in data:
            worksheet.append(row)
        
        workbook.save(self.download_path)
    
    def __str__(self) -> str:
        return f"Creating an Excel file: {self.excel_name}"
    
if __name__ == "__main__":
    SqliteToExcel(
        db_path='testing/sqlite3.db',
        download_path='testing',
        overwrite=True
    )